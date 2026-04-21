# -*- coding: utf-8 -*-
import base64
import csv
import io
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from odoo import fields, models


class InventoryExportCSV(models.TransientModel):
    _name = 'inventory.export.csv'
    _description = 'Export Inventory to CSV'

    name = fields.Char(string='File Name', default='inventory_export.zip')
    export_file = fields.Binary(string='Export File', readonly=True)
    state = fields.Selection([('choose', 'Choose'), ('get', 'Get')], default='choose')

    def _get_export_specs(self):
        return [
            {
                'filename': 'product_template.csv',
                'sheetname': 'Sản phẩm',
                'model': 'product.template',
                'fields': [
                    'id', 'name', 'type', 'uom_id',
                    'list_price', 'qty_available', 'virtual_available',
                    'need_recall', 'create_date',
                ],
            },
            {
                'filename': 'stock_warehouse.csv',
                'sheetname': 'Kho hàng',
                'model': 'stock.warehouse',
                'fields': [
                    'id', 'name', 'code',
                    'create_date',
                ],
            },
            {
                'filename': 'stock_picking_type.csv',
                'sheetname': 'Loại phiếu kho',
                'model': 'stock.picking.type',
                'fields': [
                    'id', 'name', 'code', 'warehouse_id', 'create_date', 
                ],
            },
            {
                'filename': 'stock_picking.csv',
                'sheetname': 'Phiếu kho',
                'model': 'stock.picking',
                'fields': [
                    'id', 'name', 'origin', 'partner_id', 'picking_type_id', 'state',
                    'scheduled_date', 'date_done', 'return_deadline_date', 'return_days_left',
                    'subject_id', 'create_date', 
                ],
            },
            {
                'filename': 'stock_move.csv',
                'sheetname': 'Luồng hàng',
                'model': 'stock.move',
                'fields': [
                    'id', 'name', 'reference', 'picking_id', 'product_id', 'product_uom_qty',
                    'quantity', 'product_uom', 'location_id', 'location_dest_id', 'state',
                    'date', 'create_date', 
                ],
            },
            {
                'filename': 'stock_move_line.csv',
                'sheetname': 'Chi tiết luồng hàng',
                'model': 'stock.move.line',
                'fields': [
                    'id', 'picking_id', 'move_id', 'product_id', 'lot_id', 'owner_id',
                    'package_id', 'result_package_id', 'location_id', 'location_dest_id',
                    'quantity', 'state', 'create_date', 
                ],
            },
            {
                'filename': 'res_partner.csv',
                'sheetname': 'Đối tác',
                'model': 'res.partner',
                'fields': [
                    'id', 'name', 'company_type', 'parent_id', 'email', 'phone',
                    'street', 'city', 'country_id', 'is_school', 'create_date',
                ],
            },
            {
                'filename': 'school_partner.csv',
                'sheetname': 'Trường học',
                'model': 'res.partner',
                'domain': [('is_school', '=', True)],
                'fields': [
                    'id', 'name', 'school_type', 'address', 'district', 'province', 'region',
                    'sale_user_ids', 'create_date', 
                ],
            },
            {
                'filename': 'school_class.csv',
                'sheetname': 'Lớp học',
                'model': 'res.partner.class',
                'fields': [
                    'id', 'grade', 'class_count', 'school_id', 'subject_ids', 'region',
                    'create_date', 
                ],
            },
            {
                'filename': 'school_subject.csv',
                'sheetname': 'Môn học',
                'model': 'res.partner.subject',
                'fields': [
                    'id', 'name', 'class_ids', 'supply_product_ids', 'create_date', 
                ],
            },
        ]

    def _serialize_value(self, record, field_name):
        if field_name not in record._fields:
            return ''

        field = record._fields[field_name]
        value = record[field_name]

        if field.type == 'many2one':
            return value.display_name if value else ''
        if field.type in ('many2many', 'one2many'):
            return ', '.join(value.mapped('display_name'))
        if field.type == 'boolean':
            return '1' if value else '0'
        if field.type in ('date', 'datetime'):
            return value or ''
        return value if value not in (False, None) else ''

    def _build_csv_content(self, model_name, field_names, domain=None):
        records = self.env[model_name].search(domain or [])
        output = io.StringIO()
        writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(field_names)
        for record in records:
            writer.writerow([self._serialize_value(record, field_name) for field_name in field_names])
        content = output.getvalue()
        output.close()
        return content

    def _get_field_labels(self, model_name, field_names):
        model = self.env[model_name].with_context(lang=self.env.user.lang)
        fields_info = model.fields_get(field_names, ['string'])
        labels = []
        for field_name in field_names:
            if field_name in fields_info:
                labels.append(fields_info[field_name]['string'] or field_name)
            else:
                labels.append(field_name)
        return labels

    def _build_excel_content(self, specs):
        workbook = openpyxl.Workbook()
        for spec in specs:
            sheet_title = spec.get('sheetname') or spec['filename'].replace('.csv', '')
            sheet = workbook.create_sheet(title=sheet_title[:31])
            records = self.env[spec['model']].search(spec.get('domain', []))
            fields = spec['fields']

            # Get translated field labels
            field_labels = self._get_field_labels(spec['model'], fields)

            # Write header row with translated labels
            for col_num, field_label in enumerate(field_labels, start=1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = field_label

            # Write data rows
            for row_num, record in enumerate(records, start=2):
                for col_num, field_name in enumerate(fields, start=1):
                    col_letter = get_column_letter(col_num)
                    sheet[f"{col_letter}{row_num}"] = self._serialize_value(record, field_name)

            # Define table range and add table
            table_range = f"A1:{get_column_letter(len(fields))}{len(records) + 1}"
            table = Table(displayName=spec['filename'].replace('.csv', '_table'), ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            sheet.add_table(table)

        # Remove default sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def action_export_csv(self):
        self.ensure_one()

        folder_name = 'inventory_export'
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for spec in self._get_export_specs():
                csv_content = self._build_csv_content(
                    spec['model'],
                    spec['fields'],
                    domain=spec.get('domain'),
                )
                zip_file.writestr(
                    '%s/%s' % (folder_name, spec['filename']),
                    csv_content.encode('utf-8-sig'),
                )

        self.write({
            'export_file': base64.b64encode(zip_buffer.getvalue()),
            'state': 'get',
            'name': '%s.zip' % folder_name,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'inventory.export.csv',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_export_excel(self):
        self.ensure_one()

        specs = self._get_export_specs()
        excel_content = self._build_excel_content(specs)

        self.write({
            'export_file': base64.b64encode(excel_content),
            'state': 'get',
            'name': 'inventory_export.xlsx',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'inventory.export.csv',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
